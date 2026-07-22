# Script to generate API test report in Excel.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (method, endpoint, description, status_code, pass_fail, reason)
results = [
    ("GET", "/health", "Health check", 200, "PASS", ""),
    ("GET", "/stats", "Service statistics", 200, "PASS", ""),
    ("POST", "/api/v1/auth/register", "Register user", 201, "PASS", ""),
    ("POST", "/api/v1/auth/login", "Login user", 200, "PASS", ""),
    ("GET", "/api/v1/auth/me", "Get current user", 200, "PASS", ""),
    ("GET", "/api/v1/auth/me", "Invalid token", 401, "PASS", ""),
    ("POST", "/api/v1/posts", "Create post", 201, "PASS", ""),
    ("GET", "/api/v1/posts", "List posts", 200, "PASS", ""),
    ("GET", "/api/v1/posts/{post_id}", "Get post", 200, "PASS", ""),
    ("PUT", "/api/v1/posts/{post_id}", "Update post", 200, "PASS", ""),
    ("DELETE", "/api/v1/posts/{post_id}", "Delete post", 204, "PASS", ""),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Test Report"
hf = Font(bold=True, color="FFFFFF", size=11)
hbg = PatternFill("solid", fgColor="2F5496")
pg = PatternFill("solid", fgColor="C6EFCE")
fr = PatternFill("solid", fgColor="FFC7CE")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
t = Side(style="thin")
bdr = Border(left=t, right=t, top=t, bottom=t)
for c, h in enumerate(["#", "Method", "Endpoint", "Description", "Status Code", "Pass/Fail", "Reason"], 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hbg
    cell.alignment = ctr
    cell.border = bdr
for row, (m, ep, desc, code, pf, rsn) in enumerate(results, 2):
    bg = pg if pf == "PASS" else fr
    for c, (v, a) in enumerate(
        zip([row - 1, m, ep, desc, code, pf, rsn], [ctr, ctr, lft, lft, ctr, ctr, lft]), 1
    ):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = bg
        cell.alignment = a
        cell.border = bdr
        if c == 6:
            cell.font = Font(bold=True, color="375623" if pf == "PASS" else "9C0006")
for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
wb.save("api_test_report.xlsx")
print("Saved: api_test_report.xlsx")
